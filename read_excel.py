import openpyxl as excel
import os


class ReadExcel(object):

    sheet_obj = None
    max_row = None
    max_col = None
    has_title = True

    def __init__(self, sheet_name="Algo Test.xlsx", has_title=True):
        dir_name = os.path.dirname(__file__)
        path = os.path.join(dir_name, "data_files/%s"%sheet_name)
        wb_obj = excel.load_workbook(path)
        self.sheet_obj = wb_obj.active
        self.max_row = self.sheet_obj.max_row
        self.max_col = self.sheet_obj.max_column
        self.has_title = has_title

    def get_sheet(self):
        return self.sheet_obj

    def get_has_title(self):
        return self.has_title

    def get_total_row(self):
        return self.max_row

    def get_total_column(self):
        return self.max_col

    def get_column_data(self, column_number, has_title=True):
        if column_number <= 0:
            return "Column number should be greater than or equal to 1"
        elif column_number > self.max_col:
            return "Column number exceeds total number of columns"
        else:
            data = {"header": "", "data": []}
            range_start = 1
            if has_title:
                cell_obj = self.sheet_obj.cell(row=1, column=column_number)
                data["header"] = str(cell_obj.value) if cell_obj.value else ""
                range_start = 2
            for row_number in xrange(range_start, self.max_row+1):
                cell_obj = self.sheet_obj.cell(row=row_number, column=column_number)
                data["data"].append(cell_obj.value if cell_obj.value else 0)
            data["data"] = tuple(data["data"])
            return data

    def get_row_data(self, row_number):
        if row_number <= 0:
            return "Row number should be greater than or equal to 1"
        elif row_number > self.max_row:
            return "Row number exceeds total number of rows"
        else:
            data = []
            range_start = 1
            for col_number in xrange(range_start, self.max_col+1):
                cell_obj = self.sheet_obj.cell(row=row_number, column=col_number)
                data.append(cell_obj.value if cell_obj.value else 0)
            return tuple(data)

    def get_cell_data(self, row_number, column_number):
        if row_number <= 0:
            return "Row number should be greater than or equal to 1"
        elif row_number > self.max_row:
            return "Row number exceeds total number of rows"
        if column_number <= 0:
            return "Column number should be greater than or equal to 1"
        elif column_number > self.max_col:
            return "Column number exceeds total number of columns"
        cell_obj = self.sheet_obj.cell(row=row_number, column=column_number)
        data = cell_obj.value if cell_obj.value else 0
        return data
