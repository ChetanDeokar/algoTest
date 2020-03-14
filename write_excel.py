import os
import openpyxl as xl


class WriteExcel(object):

    file_name = None
    wb = None
    sheet = None
    file_path = None

    def __init__(self, file_name='result.xlsx'):
        self.file_name = file_name
        dir_name = os.path.dirname(__file__)
        self.file_path = os.path.join(dir_name, "data_files/%s" % file_name)
        wb = xl.Workbook()
        wb.save(self.file_path)
        self.wb = wb
        self.sheet = self.wb.active

    def get_workbook(self):
        return self.wb

    def get_active_sheet(self):
        return self.sheet

    def write_row(self, data, row_number=1, column_number=1):
        for index, info in enumerate(data, start=0):
            cell = self.sheet.cell(row=row_number, column=column_number+index)
            cell.value = info
        self.wb.save(self.file_path)

    def write_column(self, data, column_number=1, row_number=1):
        for index, info in enumerate(data, start=0):
            cell = self.sheet.cell(row=index+row_number, column=column_number)
            cell.value = info

    def save(self):
        self.wb.save(self.file_path)
